from flask import Flask, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import unicodedata
import re

app = Flask(__name__)
CORS(app)

# ---------------------------------------------------------
# NAME CLEANER — removes invisible Excel characters
# ---------------------------------------------------------
def clean_name(value):
    if not value:
        return ""

    s = str(value).upper()
    s = unicodedata.normalize("NFKD", s)

    s = s.replace("\u00A0", " ")
    s = s.replace("\u200B", "")
    s = s.replace("\uFEFF", "")
    s = s.replace("\t", " ")
    s = s.replace("\n", " ")

    s = re.sub(r"[^A-Z]", "", s)

    return s.strip()


# ---------------------------------------------------------
# ---------- ROUND 1 ----------
# ---------------------------------------------------------
@app.route("/api/round1")
def round1():
    try:
        wb = load_workbook("leaderboard.xlsx", data_only=True)
        ws = wb["Round1"]

        standings = [
            ["" if c is None else c for c in r]
            for r in ws.iter_rows(
                min_row=6, max_row=14, min_col=1, max_col=4, values_only=True
            )
        ]

        strokes = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=22, max_row=33, min_col=1, max_col=23, values_only=True
            )
        ]

        points = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=38, max_row=49, min_col=1, max_col=23, values_only=True
            )
        ]

        return jsonify(
            {"status": "success", "standings": standings, "strokes": strokes, "points": points}
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---------------------------------------------------------
# ---------- ROUND 2 ----------
# ---------------------------------------------------------
@app.route("/api/round2")
def round2():
    try:
        wb = load_workbook("leaderboard.xlsx", data_only=True)
        ws = wb["Round2"]

        standings = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=6, max_row=14, min_col=1, max_col=4, values_only=True
            )
        ]

        strokes = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=22, max_row=33, min_col=1, max_col=23, values_only=True
            )
        ]

        points = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=38, max_row=49, min_col=1, max_col=23, values_only=True
            )
        ]

        # Matchplay display table
        matchplay = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=54, max_row=65, min_col=1, max_col=22, values_only=True
            )
        ]

        # Helper table for colour scoring
        HOLE_COUNT = 18
        FIRST_HOLE_COL = 5   # Column E = hole 1
        PLAYER_COL = 1       # Column A

        score_map = {}

        for row in ws.iter_rows(
            min_row=99,
            max_row=106,
            min_col=1,
            max_col=FIRST_HOLE_COL + HOLE_COUNT - 1,
            values_only=True,
        ):
            raw_name = row[PLAYER_COL - 1]
            player = clean_name(raw_name)

            if not player:
                continue

            hole_scores = {}
            for hole_idx in range(1, HOLE_COUNT + 1):
                col_index = FIRST_HOLE_COL - 1 + (hole_idx - 1)
                val = row[col_index] if col_index < len(row) else None
                try:
                    hole_scores[hole_idx] = int(val) if val is not None else 0
                except Exception:
                    hole_scores[hole_idx] = 0

            score_map[player] = hole_scores

        matchplay_colors = []
        cleaned_keys = {clean_name(k): k for k in score_map.keys()}

        for excel_row_idx, row in enumerate(
            ws.iter_rows(min_row=54, max_row=65, min_col=1, max_col=22, values_only=True),
            start=54,
        ):
            color_row = []

            for excel_col_idx, cell in enumerate(row, start=1):
                color = ""

                # Only colour player rows + hole columns
                if 58 <= excel_row_idx <= 65 and 5 <= excel_col_idx <= 22:
                    display_name = clean_name(cell)

                    if display_name:
                        hole_idx = excel_col_idx - 4  # E => 1

                        matched = cleaned_keys.get(display_name)

                        if matched:
                            score = score_map[matched].get(hole_idx, 0)

                            if score == 3:
                                color = "gold"
                            elif score == 2:
                                color = "silver"
                            elif score == 1:
                                color = "bronze"

                color_row.append(color)

            matchplay_colors.append(color_row)

        # Overall standings after 2 rounds (now includes TOTAL column)
        overall_after_2 = read_overall_table(ws, round_count=2)

        return jsonify(
            {
                "status": "success",
                "standings": standings,
                "strokes": strokes,
                "points": points,
                "matchplay": matchplay,
                "matchplay_colors": matchplay_colors,
                "overall": overall_after_2,
            }
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---------------------------------------------------------
# ---------- ROUND 3 ----------
# ---------------------------------------------------------
@app.route("/api/round3")
def round3():
    try:
        wb = load_workbook("leaderboard.xlsx", data_only=True)
        ws = wb["Round3"]

        standings = [
            ["" if c is None else c for c in r]
            for r in ws.iter_rows(
                min_row=6, max_row=14, min_col=1, max_col=4, values_only=True
            )
        ]

        strokes = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=22, max_row=33, min_col=1, max_col=23, values_only=True
            )
        ]

        points = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=38, max_row=49, min_col=1, max_col=23, values_only=True
            )
        ]

        overall_after_3 = read_overall_table(ws, round_count=3)

        return jsonify(
            {
                "status": "success",
                "standings": standings,
                "strokes": strokes,
                "points": points,
                "overall": overall_after_3,
            }
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---------------------------------------------------------
# ---------- ROUND 4 ----------
# ---------------------------------------------------------
@app.route("/api/round4")
def round4():
    try:
        wb = load_workbook("leaderboard.xlsx", data_only=True)
        ws = wb["Round4"]

        standings = [
            ["" if c is None else c for c in r]
            for r in ws.iter_rows(
                min_row=6, max_row=14, min_col=1, max_col=4, values_only=True
            )
        ]

        strokes = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=22, max_row=33, min_col=1, max_col=23, values_only=True
            )
        ]

        points = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=38, max_row=49, min_col=1, max_col=23, values_only=True
            )
        ]

        overall_after_4 = read_overall_table(ws, round_count=4)

        return jsonify(
            {
                "status": "success",
                "standings": standings,
                "strokes": strokes,
                "points": points,
                "overall": overall_after_4,
            }
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---------------------------------------------------------
# ---------- ROUND 5  (same as ROUND 2) -------------------
# ---------------------------------------------------------
@app.route("/api/round5")
def round5():
    try:
        wb = load_workbook("leaderboard.xlsx", data_only=True)
        ws = wb["Round5"]  # must match sheet name exactly

        standings = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=6, max_row=14, min_col=1, max_col=4, values_only=True
            )
        ]

        strokes = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=22, max_row=33, min_col=1, max_col=23, values_only=True
            )
        ]

        points = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=38, max_row=49, min_col=1, max_col=23, values_only=True
            )
        ]

        matchplay = [
            [cell if cell is not None else "" for cell in row]
            for row in ws.iter_rows(
                min_row=54, max_row=65, min_col=1, max_col=22, values_only=True
            )
        ]

        HOLE_COUNT = 18
        FIRST_HOLE_COL = 5
        PLAYER_COL = 1

        score_map = {}

        for row in ws.iter_rows(
            min_row=99,
            max_row=106,
            min_col=1,
            max_col=FIRST_HOLE_COL + HOLE_COUNT - 1,
            values_only=True,
        ):
            raw_name = row[PLAYER_COL - 1]
            player = clean_name(raw_name)

            if not player:
                continue

            hole_scores = {}
            for hole_idx in range(1, HOLE_COUNT + 1):
                col_index = FIRST_HOLE_COL - 1 + (hole_idx - 1)
                val = row[col_index] if col_index < len(row) else None
                try:
                    hole_scores[hole_idx] = int(val) if val is not None else 0
                except Exception:
                    hole_scores[hole_idx] = 0

            score_map[player] = hole_scores

        matchplay_colors = []
        cleaned_keys = {clean_name(k): k for k in score_map.keys()}

        for excel_row_idx, row in enumerate(
            ws.iter_rows(min_row=54, max_row=65, min_col=1, max_col=22, values_only=True),
            start=54,
        ):
            color_row = []
            for excel_col_idx, cell in enumerate(row, start=1):
                color = ""

                if 58 <= excel_row_idx <= 65 and 5 <= excel_col_idx <= 22:
                    display_name = clean_name(cell)
                    if display_name:
                        hole_idx = excel_col_idx - 4
                        matched = cleaned_keys.get(display_name)
                        if matched:
                            score = score_map[matched].get(hole_idx, 0)
                            if score == 3:
                                color = "gold"
                            elif score == 2:
                                color = "silver"
                            elif score == 1:
                                color = "bronze"

                color_row.append(color)

            matchplay_colors.append(color_row)

        overall_after_5 = read_overall_table(ws, round_count=5)

        return jsonify(
            {
                "status": "success",
                "standings": standings,
                "strokes": strokes,
                "points": points,
                "matchplay": matchplay,
                "matchplay_colors": matchplay_colors,
                "overall": overall_after_5,
            }
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---------------------------------------------------------
# ---------- ROUND 6 (same as ROUND 1, Net Scores) --------
# ---------------------------------------------------------
@app.route("/api/round6")
def round6():
    try:
        wb = load_workbook("leaderboard.xlsx", data_only=True)
        ws = wb["Round6"]

        standings = [
            ["" if c is None else c for c in r]
            for r in ws.iter_rows(
                min_row=6, max_row=14, min_col=1, max_col=4, values_only=True
            )
        ]

        strokes = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=22, max_row=33, min_col=1, max_col=23, values_only=True
            )
        ]

        points = [
            [c if c is not None else "" for c in r]
            for r in ws.iter_rows(
                min_row=38, max_row=49, min_col=1, max_col=23, values_only=True
            )
        ]

        overall_after_6 = read_overall_table(ws, round_count=6)

        return jsonify(
            {
                "status": "success",
                "standings": standings,
                "strokes": strokes,
                "points": points,
                "overall": overall_after_6,
            }
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


# ---------------------------------------------------------
# Helper: read "After X Rounds – Overall Table"
# ---------------------------------------------------------
def read_overall_table(ws, round_count):
    """
    Read the 'After X Rounds' overall table.

    Layout (all sheets Round2–Round6):

    Row 72 : title (ignored here)
    Row 73 : header row (position / arrow / player / round cols / TOTAL)
    Rows 74–81 : data rows (8 players)

    round_count controls how many round score columns we read.
    """
    # A,B,C + round_count round columns (D..)
    # +1 extra column for the new TOTAL column
    max_col = 4 + round_count

    overall = [
        [cell if cell is not None else "" for cell in row]
        for row in ws.iter_rows(
            min_row=73,  # include header row
            max_row=81,  # 8 players
            min_col=1,   # column A
            max_col=max_col,
            values_only=True,
        )
    ]
    return overall

# ---------------------------------------------------------
# RUN SERVER (Deployment-ready)
# ---------------------------------------------------------
if __name__ == "__main__":
    from os import environ

    port = int(environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
